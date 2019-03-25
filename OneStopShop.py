import openpyxl
import argparse

# TODO: Can I rewrite this without openpyxl and just use the .csv?

def startup():

    filename, workbook = None, None
    args = produce_parser()

    try:
        workbook = openpyxl.load_workbook(args.filename)
    except openpyxl.utils.exceptions.InvalidFileException:
        print("This program accepts only .xlsx, .xlsm, .xltx, and .xltm\nTry a different file.")
        return
    except FileNotFoundError:
        print("This file could not be opened. Try a different file.")
        return

    # This first sheet will always exist, since a workbook cannot exist without a sheet
    aggregated_donations = analyze(workbook[workbook.sheetnames[0]])

    if aggregated_donations is None:
        print("Either this file contains no data, or something unexpected went wrong.")
        return

    save_result(filename, workbook, aggregated_donations)


def produce_parser():
    parser = argparse.ArgumentParser(description="This tool is designed to simplify the process of analyzing and "
                                                 "interpreting the data found in FEC donation information by "
                                                 "aggregating donations for each person listed in the given Excel "
                                                 "spreadsheet.")
    parser.add_argument("filename", help="the file to read donation data from")
    parser.add_argument("-n", "--name", default="N", metavar="", help="the column where the donors name can be found")
    parser.add_argument("-c", "--committee", default="B", metavar="", help="the column where the committee "
                                                                           "receiving the donation can be found")
    parser.add_argument("-d", "--donation", default="AH", metavar="", help="the column where the amount donated "
                                                                           "can be found")
    return parser.parse_args()

def analyze(data_sheet):

    # Check the sheet format
    # TODO: Use arguments with defaults (argparse) to determine which cols should have what data

    if data_sheet["N1"].value != "contributor_name" or data_sheet["B1"].value != "committee_name" or \
            data_sheet["AH1"].value != "contribution_receipt_amount":
        print("This file is improperly formatted. Check the file and try again.")
        return

    aggregated_donations = {}

    # Note: this does not account for typos and misspelled names
    for index in range(data_sheet.min_row + 1, data_sheet.max_row + 1):

        name, org, don_amt = parse_row(data_sheet[index])

        # TODO: This is my entire data structure here: worth optimizing once it works
        # Ignore all rows that have either no org, no name, or no donation
        if name and org and don_amt:
            if name not in aggregated_donations:
                aggregated_donations[name] = {}

            if org not in aggregated_donations[name]:
                aggregated_donations[name][org] = 0

            aggregated_donations[name][org] += don_amt

    return aggregated_donations


def parse_row(row):
    # (name, organization, donation amount)
    # The indexes have to be done like this because openpyxl returns a tuple of columns
    return row[ord("N") % 65].value, row[ord("B") % 65].value, row[26 + ord("H") % 65].value


def save_result(filename, workbook, aggregated_donations):
    # TODO: This will always create a new sheet - is this the intended behavior?
    result_sheet = workbook.create_sheet("aggregate_data")

    result_sheet["A1"] = "donor_name"
    result_sheet["B1"] = "committee_name"
    result_sheet["C1"] = "aggregate_amount"

    # Start at 2, to account for the Excel double offset (start at 1 and a header row)
    curr_row = 2

    for (name, donation_entry) in aggregated_donations.items():
        for org in donation_entry:

            # Output goes to rows A, B, C
            result_sheet["A{0}".format(curr_row)] = name
            result_sheet["B{0}".format(curr_row)] = org
            result_sheet["C{0}".format(curr_row)] = donation_entry[org]
            curr_row += 1

    print("Saving result...")
    workbook.save(filename)
    return


if __name__ == '__main__':
    startup()
