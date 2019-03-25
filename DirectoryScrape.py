import requests
import re
from bs4 import BeautifulSoup
from openpyxl import load_workbook

DIRECTORY_URL = "https://directory.tufts.edu/searchresults.cgi"
WORKBOOK_NAME = "FullDataSet.xlsx"
NAME_SHEET = "UniqueNames"

def getDirectoryPage(name):
    rawPage = requests.post(DIRECTORY_URL, data={"type": "Faculty", "search": name})
    soup = BeautifulSoup(rawPage.text, "html.parser")
    for child in soup.find_all(href=re.compile("department.cgi")):
        return child.contents[0].strip()
# Modernize to not use the deprecated get_sheet_by_name
dataBook = load_workbook(WORKBOOK_NAME)
nameSheet = dataBook.get_sheet_by_name(NAME_SHEET)

for index in range(2, nameSheet.max_row + 1):
    currentName = nameSheet["A{0}".format(index)].value
    affiliation = getDirectoryPage(currentName)
    nameSheet["B{0}".format(index)] = affiliation

    # Rudimentary backup every 100 entries
    if index % 100 == 0:
        print("Progress: {:.2%}".format(index / (nameSheet.max_row + 1)))
        dataBook.save(WORKBOOK_NAME)

dataBook.save(WORKBOOK_NAME)
