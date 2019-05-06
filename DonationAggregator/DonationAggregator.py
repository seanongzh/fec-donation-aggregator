import openpyxl
import argparse
import requests
import os

# TODO: Include support for csv files using the built in Python CSV parser

FEC_API = "VDkmeFlFlO9ZRao7AyDyPMrgEeSdwJXO8UdN7faS"
BASE_URL = "http://api.open.fec.gov/v1/committee/"

### Main program flow ###

def startup():

    args = produce_parser()

    # Create a master workbook to save results into
    master_book = openpyxl.Workbook()
    master_donations = {}

    # The given directory will contain a series of subdirectories
    for donor_name in os.listdir(args.directory):

        subdirectory = os.path.join(args.directory, donor_name)

        if os.path.isdir(subdirectory):

            # This first sheet will always exist, since a workbook cannot exist without a sheet
            aggregated_donations = analyze(subdirectory, args.committee, args.donation, args.id)

            if aggregated_donations is None:
                print("Either this file contains no data, or something unexpected went wrong.")
                return

            master_donations[donor_name] = aggregated_donations

    master_book = save_result(master_book, master_donations)
    print("Saving result...")
    master_book.save(os.path.join(args.directory, args.output))


def analyze(subdirectory, committee_col, donation_col, comm_id_col):

    aggregated_donations = {}

    # This gives us every file for each subdirectory
    for file in os.listdir(subdirectory):
        if file.endswith(".xlsx"):
            workbook = open_xlsx(os.path.join(subdirectory, file))

            if workbook is None:
                return

            data_sheet = workbook[workbook.sheetnames[0]]

            # Check the sheet format
            if data_sheet[committee_col + "1"].value != "committee_name" or \
                    data_sheet[donation_col + "1"].value != "contribution_receipt_amount" or \
                    data_sheet[comm_id_col + "1"].value != "committee_id":
                print("This file is improperly formatted. Check the file and try again.")
                print(os.path.join(subdirectory, file))
                continue

            # Note: this does not account for typos and misspelled names
            for index in range(data_sheet.min_row + 1, data_sheet.max_row + 1):

                org, don_amt, org_id = parse_row(data_sheet[index], committee_col, donation_col, comm_id_col)

                # TODO: This is my entire data structure here: worth optimizing once it works
                # Ignore all rows that have either no org, no name, or no donation
                if org and don_amt:
                    if org not in aggregated_donations:
                        aggregated_donations[org] = {}
                        aggregated_donations[org]["amount"] = 0
                        aggregated_donations[org]["id"] = org_id

                    aggregated_donations[org]["amount"] += don_amt

                # Print progress as a percent every 500 entries
                if index % 500 == 0:
                    print("Progress: {:.2%}".format(index / (data_sheet.max_row + 1)))

    return aggregated_donations


def save_result(master_book, aggregated_donations):

    result_sheet = master_book.active

    # Write headers to the results sheet
    result_sheet["A1"] = "donor_name"
    result_sheet["B1"] = "committee_name"
    result_sheet["C1"] = "committee_id"
    result_sheet["D1"] = "committee_affiliation"
    result_sheet["E1"] = "aggregate_amount"

    # Start at 2, to account for the Excel double offset (start at 1 and a header row)
    curr_row = 2

    # Memoize it, baby (160 ftw)
    committee_party = {}

    for (name, donation_entry) in aggregated_donations.items():
        for org in donation_entry:

            # Find the party affiliation of the committee
            if donation_entry[org]["id"] not in committee_party:
                committee_party[donation_entry[org]["id"]] = get_committee_party(donation_entry[org]["id"])

            # Output goes to rows A, B, C, D, E
            result_sheet["A{}".format(curr_row)] = name
            result_sheet["B{}".format(curr_row)] = org
            result_sheet["C{}".format(curr_row)] = donation_entry[org]["id"]
            result_sheet["D{}".format(curr_row)] = committee_party[donation_entry[org]["id"]]
            result_sheet["E{}".format(curr_row)] = donation_entry[org]["amount"]

            curr_row += 1

    return master_book


### Helper functions ###


def produce_parser():
    parser = argparse.ArgumentParser(description="This tool is designed to simplify the process of analyzing and "
                                                 "interpreting the data found in FEC donation information by "
                                                 "aggregating donations for each person listed in the given dataset.")
    parser.add_argument("directory", help="the directory to read donation data from")
    parser.add_argument("output", help="the filename to save the output sheet to")
    parser.add_argument("-c", "--committee", default="B", metavar="", help="the column where the committee "
                                                                           "receiving the donation can be found")
    parser.add_argument("-d", "--donation", default="AH", metavar="", help="the column where the amount donated "
                                                                           "can be found")
    parser.add_argument("-i", "--id", default="A", metavar="", help="indicates which column the committee ID can be "
                                                                    "found in. The program will determine the political"
                                                                    " affiliation of each committee using the FEC API")
    return parser.parse_args()


# Safely opens a .xlsx file using openpyxl, catching errors and providing error output along the way
def open_xlsx(filename):
    try:
        return openpyxl.load_workbook(filename)
    except openpyxl.utils.exceptions.InvalidFileException:
        print("This program accepts only .xlsx, .xlsm, .xltx, and .xltm\nTry a different file.")
        return None
    except FileNotFoundError:
        print("This file could not be opened. Try a different file.")
        return None
    except Exception as error:
        print("An unknown error occurred", error)
        return None


# Returns information from the given row as a tuple of name, organization, donation amount, and committee ID number
# (name, organization, donation amount, committee ID number)
def parse_row(row, committee_col, donation_col, comm_id_col=None):
    # The indexes have to be done like this because openpyxl returns a tuple of columns
    return row[letter_number(committee_col)].value, row[letter_number(donation_col)].value, \
           row[letter_number(comm_id_col)].value


# TODO: There is certainly a way to do this with more algorithmic ~sizzle~
# Returns the number associated with the Excel column of the given number.
# E.g. A --> 0, AC --> 28
def letter_number(col):

    if len(col) == 1:
        return ord(col) % 65
    elif len(col) == 2:
        return (((ord(col[0]) % 65) + 1) * 26) + (ord(col[1]) % 65)
    else:
        return None


# Uses the FEC API to retrieve the party affiliated with the committee with the given ID
# E.g. DEM for Democratic, REP for Republican. Returns "UNK" for committees with no registered party
def get_committee_party(id_number):
    response = requests.get(BASE_URL + id_number, params={"api_key": FEC_API})
    response_data = response.json()
    if response.status_code == 200:
        party = response_data["results"][0]["party"] if response_data["results"][0]["party"] is not None else "UNK"
        return party
    else:
        print("Unexpected response code: ", response.status_code)
        return "UNK"


if __name__ == '__main__':
    startup()
