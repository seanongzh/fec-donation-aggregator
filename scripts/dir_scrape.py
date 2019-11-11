# USAGE: dir_scrape.py [file.xlsx]
# This script works on Excel Sheets with a single column in the A column of
# a unique listing of names. It searches for every name in the tufts directory,
# then strips out a department if it is found and places it in the B column next
# to the search query.
# (Spreadsheet needs to be "master")

DIR_URL = "https://whitepages.tufts.edu/searchresults.cgi"

import openpyxl, sys, getopt, re, requests
from bs4 import BeautifulSoup

def startup(file):
    # Open file given as command line argument
    wb = openpyxl.load_workbook(file)
    ws = wb["master"]

    for i in range(2, ws.max_row + 1):
        name = ws.cell(row = i, column = 1).value
        dir_data = get_dir_data(name)

        # Department/College
        ws.cell(row = i, column = 2, value = dir_data[0])
        # Title/Major
        ws.cell(row = i, column = 3, value = dir_data[1])
        # Type (student, faculty, staff)
        ws.cell(row = i, column = 4, value = dir_data[2])

        if i % 100 == 0:
            print("Progress: {:.2%}".format(i / (ws.max_row + 1)))
            wb.save(file)

    wb.save(file)

def get_dir_data(name):
    page = requests.post(DIR_URL, data={"search": name})
    soup = BeautifulSoup(page.text, "html.parser")
    table = []
    for child in soup.find_all('td', string=True):
        table.append(child.contents[0])
    if table != []:
        print(table)
        if table[-2].strip() != "Student":
            table[2] = table[2].contents[0]
        table = [table[1].strip(), table[2].strip(), table[-2].strip()]
    else:
        table = ["UNK", "UNK", "UNK"]
    return table
    

if __name__ == '__main__':
    startup(sys.argv[1])