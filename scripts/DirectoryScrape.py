import argparse
import openpyxl
import requests
import re
from bs4 import BeautifulSoup

DIRECTORY_URL = "https://directory.tufts.edu/searchresults.cgi"
WORKBOOK_NAME = "DirectoryResults_2017-2018.xlsx"
NAME_SHEET = "DirectoryResults"

# This script works on Excel Sheets with a single column in the A column of
# a unique listing of names. It searches for every name in the tufts directory,
# then strips out a department if it is found and places it in the B column next
# to the search query.

# open_xlsx safely opens filename using openpyxl, returning the file object
def open_xlsx(filename):
    try:
        return openpyxl.load_workbook(filename)
    except openpyxl.utils.exceptions.InvalidFileException:
        print("This program accepts only .xlsx, .xlsm, .xltx, and .xltm\nTry a different file.")
        return None
    except FileNotFoundError:
        print("This file could not be opened. Try a different file.")
        return None
    except RuntimeError as error:
        print("An unknown error occurred", error)
        return None

def getDirectoryPage(name):
    rawPage = requests.post(DIRECTORY_URL, data={"type": "Faculty", "search": name})
    soup = BeautifulSoup(rawPage.text, "html.parser")
    for child in soup.find_all(href=re.compile("department.cgi")):
        return child.contents[0].strip()

# Parser for command-line arguments
parser = argparse.ArgumentParser(description='Map names in a spreadsheet to \
                                 the Tufts University directory.')
parser.add_argument('file', help='An Excel file with a unique listing of names \
                    in the first column, will be overridden')
args = parser.parse_args()

dataBook = open_xlsx(args.file)
nameSheet = dataBook[NAME_SHEET]

for index in range(2, nameSheet.max_row + 1):
    currentName = nameSheet["A{0}".format(index)].value
    affiliation = getDirectoryPage(currentName)
    nameSheet["B{0}".format(index)] = affiliation

    # Rudimentary backup every 100 entries
    if index % 100 == 0:
        print("Progress: {:.2%}".format(index / (nameSheet.max_row + 1)))
        dataBook.save(args.file)

dataBook.save(args.file)
