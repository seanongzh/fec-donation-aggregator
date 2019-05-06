from openpyxl import load_workbook

FILE_NAME = "2017-2018_filtered_typos.xlsx"
DATA_SHEET = "FilteredSourceData"
TYPO_SHEET = "TypoListing"


dataBook = load_workbook(FILE_NAME)
dataSheet = dataBook[DATA_SHEET]
typoSheet = dataBook[TYPO_SHEET]

typoList = {}

for index in range(typoSheet.min_row + 1, typoSheet.max_row + 1):
    if typoSheet["B{0}".format(index)].value is not None:
        typoList[typoSheet["A{0}".format(index)].value] = typoSheet["B{0}".format(index)].value

for index in range(dataSheet.min_row + 1, dataSheet.max_row + 1):
    first = dataSheet["Q{0}".format(index)].value
    last = dataSheet["S{0}".format(index)].value

    # Skip over names that have no first or last
    if first is None or last is None:
        continue

    currName = last + ", " + first

    if currName in typoList:
        dataSheet["O{0}".format(index)] = typoList[currName]
    else:
        dataSheet["O{0}".format(index)] = currName

dataBook.save(FILE_NAME)
