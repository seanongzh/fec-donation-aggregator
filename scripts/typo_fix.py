# USAGE: typo_fix.py [file1.xlsx] [file2.xlsx]
# Corrects some data in file1 based on a (wrong_data, correct_data) list in file2
# List is structured in file2 as (assumed to be the spreadsheet "master"):
#   [wrong_data] | [correct_data]
# Data needs to be in the first column in typo_fix sheet in file1
#   If data in file1 is found in wrong_data, data will be replaced with correct_data

import openpyxl, sys, getopt

def startup(file1, file2):
    # Open file given as command line argument
    wb_actual = openpyxl.load_workbook(file1)
    wb_fix    = openpyxl.load_workbook(file2)
    ws_actual = wb_actual["typo_fix"]
    ws_fix    = wb_fix["master"]

    # For each piece of data in the actual ws (with header row)
    for i in range(2, ws_actual.max_row + 1):
        actual = ws_actual.cell(row = i, column = 1)
        if i % 250 == 0:
            print("Fixing " + actual.value + " at row " + str(i))
        # Find the typo fix in fix ws (with header row)
        for j in range(2, ws_fix.max_row + 1): 
            typo = ws_fix.cell(row = j, column = 1)
            if typo.value == actual.value:
                fixed = ws_fix.cell(row = j, column = 2)
                ws_actual.cell(row = i, column = 1, value = fixed.value)

    wb_actual.save(file1)

if __name__ == '__main__':
    startup(sys.argv[1], sys.argv[2])