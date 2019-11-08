# USAGE: extract_earmark.py [file.xlsx]
# Extracts some earmark data from a column of contribution memos in file.xlsx
# List is structured in file as a single column (assumed to be the spreadsheet "master"):
# Relevant data needs to be in the first column and can take the following forms:
#   - "EARMARKED CONTRIBUTION..."
#   - "EARMARKED FOR..."
#   Script will replace earmark memo with a committee name/ID if it exists
#   Otherwise, memo will remain the same
#   Leading whitespace and asterisks will be removed

import openpyxl, sys, getopt, re

def startup(file):
    # Open file given as command line argument
    wb = openpyxl.load_workbook(file)
    ws = wb["master"]

    # For each string in the actual ws (with header row)
    for i in range(2, ws.max_row + 1):
        actual = ws.cell(row = i, column = 1)
        if actual.value is not None:
            if(i % 250 == 0):
                print("Extracting " + actual.value + " at row " + str(i))     
            # Remove leading whitespaces and asterisk
            strip_str = re.sub("^(\*\ )+", "", actual.value)
            id = ""
            name = ""

            # "EARMARKED FOR [COMMITTEE XX] ([ID])"
            if re.search("^(EARMARKED FOR){1}", strip_str) is not None:
                # "[COMMITTEE XX] ([ID])"
                name = re.split("^(EARMARKED FOR )", strip_str)[2]
                # "([ID])"
                if re.search("(\(){1}(.)*(\)){1}", name) is not None:
                    id = re.search("(\(){1}(.)*(\)){1}", name).group()
                name_id_list = re.split(id, name)
                # "[ID]"
                id = name_id_list[1]
                name = name_id_list[0]
                # "[COMMITTEE XX]"
                name = re.split("(\ \()$", name)[0]

            ws.cell(row = i, column = 1, value = strip_str)
            # Committee name and ID go to col 2 and 3
            ws.cell(row = i, column = 2, value = name)
            ws.cell(row = i, column = 3, value = id)

    wb.save(file)

if __name__ == '__main__':
    startup(sys.argv[1])