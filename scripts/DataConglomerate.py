from openpyxl import load_workbook


FILE_NAME = "FullDataSet.xlsx"
DONO_SHEET = "Donations"
NAME_SHEET = "UniqueNames"

class NameEntry:

    def __init__(self, affiliation):
        self.donations = {}
        self.affiliation = affiliation


dataList = {}

dataBook = load_workbook(FILE_NAME)
donoSheet = dataBook[DONO_SHEET]
nameSheet = dataBook[NAME_SHEET]

# Preload the data with names and affiliations
for index in range(nameSheet.min_row + 1, nameSheet.max_row + 1):
    # Create a new listing for each unique name, with the affiliation generated from DirectoryScrape.py

    if nameSheet["B{0}".format(index)].value is not None and nameSheet["A{0}".format(index)].value is not None:
        dataList[nameSheet["A{0}".format(index)].value] = NameEntry(nameSheet["B{0}".format(index)].value)

# Aggregate the data
for index in range(donoSheet.min_row + 1, donoSheet.max_row + 1):

    # The donoSheet should have names in A, org names in B, and amounts in C
    name = donoSheet["A{0}".format(index)].value
    org = donoSheet["B{0}".format(index)].value
    amount = donoSheet["C{0}".format(index)].value

    # Ignore people not preloaded in the datalist
    if name in dataList:
        entry = dataList[name]
        # If the org hasn't been recorded before, add it to the list of targets
        if org not in entry.donations:
            entry.donations[org] = 0

        entry.donations[org] += amount

# Output the data
# Start at 2 to catch the Excel double-offset (1-index and header row)
currRow = 2

for (name, entry) in dataList.items():
    for org in entry.donations:
        # Output goes to rows H, I, J, K
        donoSheet["H{0}".format(currRow)] = name
        donoSheet["I{0}".format(currRow)] = entry.affiliation
        donoSheet["J{0}".format(currRow)] = org
        donoSheet["K{0}".format(currRow)] = entry.donations[org]
        currRow += 1

dataBook.save(FILE_NAME)
