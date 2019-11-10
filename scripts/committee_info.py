# USAGE: committee_info.py [filename.xlsx]
# Calls FEC API for each committee in a given list, returns party/geo affiliations and committee type/designation for each.
# List is structured in a spreadsheet as:
#   [committee_name] | [committee_id] | [committee_party] | [committee_geo] | [committee_designation] | [committee_type]
# List is assumed to be the spreadsheet labelled "master" in a given workbook.

import openpyxl, requests, sys, getopt

# FEC API key is contained in config.py
import config
FEC_API = config.fec_key
BASE_URL = "http://api.open.fec.gov/v1/committee/"
NAME_URL = "http://api.open.fec.gov/v1/names/committees/"

def startup(filename):
    # Open file given as command line argument
    wb = openpyxl.load_workbook(filename)
    ws = wb["master"]

    # For each committee (each row)
    for i in range(2, ws.max_row + 1):
        # Access committee_id, held in 2nd column
        if ws.cell(row = i, column = 2).value is not None:
            id = ws.cell(row = i, column = 2).value
            print(id)
        else:
            print(ws.cell(row = i, column = 1).value)
            id = get_committee_id(ws.cell(row = i, column = 1).value)
            ws.cell(row = i, column = 2, value = id)

        # Collapse code into one function so that API calls can be reduced
        info = get_committee_info(id)

        # Write party affiliation to 3rd column
        ws.cell(row = i, column = 3, value = info[0])

        # Write geo (state) affiliation to 4th column
        ws.cell(row = i, column = 4, value = info[1])

        # Write designation to 5th column
        ws.cell(row = i, column = 5, value = info[2])

        # Write type to 6th column
        ws.cell(row = i, column = 6, value = info[3])

        if i % 50 == 0:
            wb.save(filename)

    wb.save(filename)

# Uses the FEC API to retrieve the committee ID associated with a given name
def get_committee_id(name):
    response = requests.get(NAME_URL, params={"api_key": FEC_API, "q": name})
    response_data = response.json()
    if response.status_code == 200:
        id = response_data["results"][0]["id"] if response_data["results"][0]["id"] is not None else "UNK"
        return id
    else:
        print("Unexpected response code (ID): ", response.status_code)
        return "UNK"
    
def get_committee_info(id_number):
    response = requests.get(BASE_URL + id_number, params={"api_key" : FEC_API})
    response_data = response.json()
    if response.status_code == 200:
        party = response_data["results"][0]["party"] if response_data["results"][0]["party"] is not None else "UNK"
        geo = response_data["results"][0]["state"] if response_data["results"][0]["state"] is not None else "UNK"
        desgn = response_data["results"][0]["designation_full"] if response_data["results"][0]["designation_full"] is not None else "UNK"
        type = response_data["results"][0]["committee_type_full"] if response_data["results"][0]["committee_type_full"] is not None else "UNK"
        return [party, geo, desgn, type]
    else:
        print("Unexpected response code: ", response.status_code)
        return ["UNK", "UNK", "UNK", "UNK"]

# Uses the FEC API to retrieve the party affiliated with the committee with the given ID
# E.g. DEM for Democratic, REP for Republican. Returns "UNK" for committees with no registered party
def get_committee_party(id_number):
    response = requests.get(BASE_URL + id_number, params={"api_key": FEC_API})
    response_data = response.json()
    if response.status_code == 200:
        party = response_data["results"][0]["party"] if response_data["results"][0]["party"] is not None else "UNK"
        return party
    else:
        print("Unexpected response code (Party): ", response.status_code)
        return "UNK"

# Uses the FEC API to retrieve the geographic area (state) affiliated with the committee with the given ID
# (Standard two-letter state codes)
def get_committee_geo(id_number):
    response = requests.get(BASE_URL + id_number, params={"api_key": FEC_API})
    response_data = response.json()
    if response.status_code == 200:
        geo = response_data["results"][0]["state"] if response_data["results"][0]["state"] is not None else "UNK"
        return geo
    else:
        print("Unexpected response code (geo): ", response.status_code)
        return "UNK"

# Uses the FEC API to retrieve the designation of the committee with the given ID
# (A = authorized by a candidate
#  J = joint fundraising committee
#  P = principal campaign committee of a candidate
#  U = unauthorized
#  B = lobbyist/registrant PAC
#  D = leadership PAC)
def get_committee_designation(id_number):
    response = requests.get(BASE_URL + id_number, params={"api_key": FEC_API})
    response_data = response.json()
    if response.status_code == 200:
        desgn = response_data["results"][0]["designation_full"] if response_data["results"][0]["designation_full"] is not None else "UNK"
        return desgn
    else:
        print("Unexpected response code (des): ", response.status_code)
        return "UNK"

# Uses the FEC API to retrieve the type of the committee with the given ID
# (- C communication cost
#  - D delegate
#  - E electioneering communication
#  - H House
#  - I independent expenditor (person or group)
#  - N PAC - nonqualified
#  - O independent expenditure-only (super PACs)
#  - P presidential
#  - Q PAC - qualified
#  - S Senate
#  - U single candidate independent expenditure
#  - V PAC with non-contribution account, nonqualified
#  - W PAC with non-contribution account, qualified
#  - X party, nonqualified
#  - Y party, qualified
#  - Z national party non-federal account)
def get_committee_type(id_number):
    response = requests.get(BASE_URL + id_number, params={"api_key": FEC_API})
    response_data = response.json()
    if response.status_code == 200:
        type = response_data["results"][0]["committee_type_full"] if response_data["results"][0]["committee_type_full"] is not None else "UNK"
        return type
    else:
        print("Unexpected response code (type): ", response.status_code)
        return "UNK"

if __name__ == '__main__':
    startup(sys.argv[1])