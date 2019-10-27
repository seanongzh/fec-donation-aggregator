# USAGE: party_geo.py [filename.xlsx]
# Calls FEC API for each committee in a given list, returns party and geo affiliations for each.
# List is structured in a spreadsheet as:
#   [committee_name] | [committee_id] | [committee_party] | [committee_geo] | [committee_designation]
# List is assumed to be the spreadsheet labelled "master" in a given workbook.

import openpyxl, requests, sys, getopt

# FEC API key is contained in config.py
import config
FEC_API = config.fec_key
BASE_URL = "http://api.open.fec.gov/v1/committee/"

def startup(filename):
    # Open file given as command line argument
    wb = openpyxl.load_workbook(filename)
    ws = wb["master"]

    # For each committee (each row)
    for i in range(2, ws.max_row + 1):
        # Access committee_id, held in 2nd column
        id = ws.cell(row = i, column = 2).value

        # Write party affiliation to 3rd column
        ws.cell(row = i, column = 3, value = get_committee_party(id))

        # Write geo (state) affiliation to 4th column
        ws.cell(row = i, column = 4, value = get_committee_geo(id))

        # Write designation to 5th column
        ws.cell(row = i, column = 5, value = get_committee_designation(id))
    
    wb.save(filename)

# Uses the FEC API to retrieve the party affiliated with the committee with the given ID
# E.g. DEM for Democratic, REP for Republican. Returns "UNK" for committees with no registered party
def get_committee_party(id_number):
    response = requests.get(BASE_URL + id_number, params={"api_key": FEC_API})
    response_data = response.json()
    if response.status_code == 200:
        party = response_data["results"][0]["party"] if response_data["results"][0]["party"] is not None else "UNK"
        return party
    else:
        print("Unexpected response code: ", response.status_code)
        return "UNK"

# Uses the FEC API to retrieve the geographic area (state) affiliated with the committee with the given ID
# (Standard two-letter state codes)
def get_committee_geo(id_number):
    response = requests.get(BASE_URL + id_number, params={"api_key": FEC_API})
    response_data = response.json()
    if response.status_code == 200:
        geo = response_data["results"][0]["state"] if response_data["results"][0]["state"] is not None else "UNK"
        return geo
    else:
        print("Unexpected response code: ", response.status_code)
        return "UNK"

# Uses the FEC API to retrieve the designation of the committee with the given ID
# (A = authorized by a candidate
#  J = joint fundraising committee
#  P = principal campaign committee of a candidate
#  U = unauthorized
#  B = lobbyist/registrant PAC
#  D = leadership PAC)
def get_committee_designation(id_number):
    response = requests.get(BASE_URL + id_number, params={"api_key": FEC_API})
    response_data = response.json()
    if response.status_code == 200:
        geo = response_data["results"][0]["designation"] if response_data["results"][0]["designation"] is not None else "UNK"
        return geo
    else:
        print("Unexpected response code: ", response.status_code)
        return "UNK"

if __name__ == '__main__':
    startup(sys.argv[1])