# USAGE: copy_info.py [file.xlsx] [from_s] [from_e] [to_s] [to_e]
# Corrects some data from col [from_s] to [from_e] to col [to_s] to [to_e] in [file.xlsx]
#   [from_s]-[from_e] needs to be the same range as [to_s]-[to_e]

import openpyxl, sys, getopt

def startup(file, from_s, from_e, to_s, to_e):
    # Open file given as command line argument
    wb = openpyxl.load_workbook(file)
    ws = wb["master"]
    # Check if cols are the same range
    if (from_e - from_s) == (to_e - to_s):
        # For each row in the ws (with header row)
        for i in range(2, ws.max_row + 1):
            # If dest col is empty (no override), copy contents from source to dest
            if ws.cell(row = i, column = to_s).value is None:
            # For each col in source range
                for j in range(0, from_e - from_s + 1):
                    ws.cell(row = i, column = to_s + j, value = ws.cell(row = i, column = from_s + j).value)
                if i % 250 == 0:
                    print("Copying at row " + str(i))
    wb.save(file)

if __name__ == '__main__':
    startup(sys.argv[1], int(sys.argv[2]), int(sys.argv[3]), int(sys.argv[4]), int(sys.argv[5]))